import serial
from openpyxl import Workbook
import statistics

port = "/dev/tty.wchusbserial110"  # Replace with the correct port name
column_index = 1  # Set the column index (e.g., 1 for column A)
start_row = 2  # Set the starting row number
i = 0
flag=1
number = 1;
workbook = Workbook()
sensor = []
worksheet = workbook.active
labelling = ["Ground Truth","Angle", "Laser", "Maxbotix Mean", "Std Deviantion", "Variance"]
for x in range(0, len(labelling)):
    worksheet.cell(row=1, column=column_index+x, value=labelling[x])
ruler_tape = int(input("Enter the tape reading: "))
while True: 
    if flag==1:
        sense = int(input("Place the sensor and press 1 or 0 to exit: "))
        if sense == 1:
            angle =int(input("Enter the angle: "))
            laser = int(input("Enter the laser reading: "))
            angle = angle-91
            ruler_tape = ruler_tape + 1100 + 104
            laser = laser -214 + 104
            #worksheet.cell(row=start_row-1, column=column_index+1, value=("Testing "+ str(number) +" at: "  + str(ruler_tape) + "mm" +" with angle: "+ str(angle)))
            uart = serial.Serial(port, 9600, timeout=1)
            while(i<100):              
                sensorValue = uart.read(12)
                dat = sensorValue[1:5]
                dat = int(dat,10)
                dat = int(dat)
                sensor.append(dat)
                #print(dat)        
                i+=1     
            uart.close()
            #print(sensor)
            mean = statistics.mean(sensor)
            variance = statistics.variance(sensor)
            standard_deviation = statistics.stdev(sensor)
            data = [ruler_tape,angle,laser, mean, standard_deviation, variance]
            for x in range(0,len(data)):
                worksheet.cell(row=start_row, column=column_index+x, value=data[x])
            number+=1
            start_row += 1
            workbook.save("Angle Data Sensor 7-2.xlsx")
            sensor =[]
        elif sense == 0:
            print("Thank you")
            flag=0
        else:
            print("Wrong input")
        i=0
        sense=0
        #workbook.save("Angle experiment2-.xlsx")
