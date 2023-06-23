## originally written by : Prafull Moona
## modified by : Ashok 

import serial
from openpyxl import Workbook
import statistics

port = "/dev/tty.wchusbserial1120"  # Replace with the correct port name
uart = serial.Serial(port, 9600, timeout=1)

######### workbook setup variables ############
column_index = 1  # Set the column index (e.g., 1 for column A)
start_row = 2  # Set the starting row number
row_100 =2
col_100 =10
i = 0
flag=1
number = 1

#### setting up the excel workbook #####
workbook = Workbook() 
worksheet = workbook.active
labelling = ["Ground Truth", "Laser", "Maxbotix Mean", "Std Deviantion", "Variance", "Error", "Percentage Error"]

for x in range(0, len(labelling)):
    worksheet.cell(row=1, column=column_index+x, value=labelling[x])



sensor_readings = []


while (flag == 1): 
    sense = int(input("Place the sensor and press 1 or 0 to exit: "))
    if sense == 1:
        #angle =int(input("Enter the angle: "))
        laser = int(input("Enter the laser reading: "))
        ruler_tape = int(input("Enter the tape reading: "))
        angle=90
        angle = angle-90
        ruler_tape = ruler_tape + 1100
        laser = laser -214
        
        #worksheet.cell(row=start_row-1, column=column_index+1, value=("Testing "+ str(number) +" at: "  + str(ruler_tape) + "mm" +" with angle: "+ str(angle)))
        
        uart = serial.Serial(port, 9600, timeout=1)
        while(i<100):           
            sensorValue = uart.read(12)
            #uart.println(sensorValue)
            dat = sensorValue[1:5]

            try:
                dat = int(dat,10)
                #print(dat)
            except ValueError:
                #uart.close()
                continue 

            ## printing data on serial for debugging
            # lines = uart.readline()
            # print(lines)

            dat = int(dat) - 104
            #print(dat)
            sensor_readings.append(dat)
            #print(sensor_readings)
            worksheet.cell(row=row_100, column=col_100, value=dat)
            #print(dat)        
            i+=1
            row_100+=1
        uart.close()
        #print(sensor_readings)
        #print(sensor)
        col_100+=1
        row_100=2
        mean = statistics.mean(sensor_readings)
        variance = statistics.variance(sensor_readings)
        standard_deviation = statistics.stdev(sensor_readings)
        error = mean - ruler_tape
        error_percentage = (error/ruler_tape)*100
        data = [ruler_tape,laser, mean, standard_deviation, variance,error, error_percentage]

        for x in range(0,len(data)):
            worksheet.cell(row=start_row, column=column_index+x, value=data[x])
        number+=1
        start_row += 1
        workbook.save("Distance Experiment 5-2'.xlsx")
        sensor =[]

    elif sense == 0:
        print("Thank you")
        flag=0

    else:
        print("Wrong input")
        
    i=0
    sense=0
    #workbook.save("Angle experiment2-.xlsx")
